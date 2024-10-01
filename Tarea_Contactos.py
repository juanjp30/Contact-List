'''
Desarrolla un programa en Python que permita a los usuarios gestionar una lista de contactos. El programa debe cumplir con los 
siguientes requisitos:
1. AGREGAR UN CONTACTO: Permitir al usuario agregar un contacto nuevo con nombre, número de teléfono y correo electrónico.
2. MOSTRAR CONTACTOS: Mostrar todos los contactos en la lista, incluyendo su nombre, número de teléfono y correo electrónico.
3. BUSCAR UN CONTACTO: Permitir al usuario buscar un contacto por nombre.
4. ELIMINAR UN CONTACTO: Permitir al usuario eliminar un contacto de la lista por nombre.
5. GUARDAR Y CARGAR CONTACTOS: Guardar la lista de contactos en un archivo para que se mantenga entre ejecuciones del programa.

Requisitos (Obligatorios): 
* Cargar la lista de contactos desde el archivo al iniciar el programa.
* Utiliza estructuras de datos básicas como listas y diccionarios.
* Manejar la entrada y salida de archivos para la persistencia de datos.

Requisitos (Opcionales):
* Validar la entrada del usuario para asegurar que los datos proporcionados son correctos (Por ejemplo, formato del correo 
  electrónico válido).
*Implentar una interfaz de línea de comandos simple para interactuar con el sistema.

'''
import openpyxl

agenda_contactos = []

def cargar_contactos_desde_excel(nombre_archivo):
    global agenda_contactos
    agenda_contactos = []
    
    try:
        # Cargar el libro de trabajo
        wb = openpyxl.load_workbook(nombre_archivo)
        
        # Seleccionar la hoja activa
        hoja = wb.active
        
        # Iterar sobre las filas, empezando desde la segunda (ignorando los encabezados)
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            nombre, telefono, correo = fila
            contacto = {
                "Nombre": nombre,
                "Telefono": str(telefono),  # Convertir a string por si acaso
                "Correo": correo
            }
            agenda_contactos.append(contacto)
        
        print(f"Se cargaron {len(agenda_contactos)} contactos desde {nombre_archivo}")
    except FileNotFoundError:
        print(f"El archivo {nombre_archivo} no se encontró. Se iniciará con una agenda vacía.")
    except Exception as e:
        print(f"Ocurrió un error al cargar el archivo: {e}")
    
    return agenda_contactos

#Cargar contactos al inicio del programa con la funcion: cargar_contactos_desde_excel
nombre_archivo = "agenda_contactos.xlsx"
agenda_contactos = cargar_contactos_desde_excel(nombre_archivo)

#Funciones para cada una de las opciones del menú
def add_contact():
    print('\nHas elegido la opción 1: Agregar un Contacto')
    nombre = input('Ingrese el nombre del contacto: ').capitalize()
    telefono = input('Ingrese el numero de teléfono del contacto: ')
    correo = input('Ingrese el e_mail del contacto: ').capitalize()
    contacto = {
        'Nombre': nombre,
        'Telefono': telefono,
        'Correo': correo
    }
    agenda_contactos.append(contacto)

def show_contact():
    print('\nHas elegido la opción 2: Mostrar Contactos')
    if not agenda_contactos:
        print("No hay contactos guardados.")
    else:
        print("Lista de contactos:")
        for i, contacto in enumerate(agenda_contactos, 1): # enumerate() devuelve pares de (índice, elemento) comenzand desde 1.
            print(f"\nContacto {i}:")
            for clave, valor in contacto.items(): # .items() devuelve cada par clave-valor (key-value) del diccionario.
                print(f"{clave.capitalize()}: {valor}")

def find_contact():
    print('\nHas elegido la opción 3: Buscar un Contacto')
    nombre_buscado = input('Ingrese el nombre del contacto que desea buscar: ').capitalize()
    for contacto in agenda_contactos:
        if contacto['Nombre'].capitalize() == nombre_buscado:
            print(f"Nombre: {contacto['nombre'].capitalize()}")
            print(f"Teléfono: {contacto['telefono']}")
            print(f"Correo: {contacto['correo'].capitalize()}")
            return
    print('Contacto no encontrado')

def delete_contact():
    print('\nHas elegido la opción 4: Eliminar un Contacto')
    nombre_eliminado = input('Ingrese el nombre del contacto que desea eliminar: ').capitalize()
    for contacto in agenda_contactos:
        if contacto['Nombre'].capitalize() == nombre_eliminado:
            agenda_contactos.remove(contacto)
            print('Contacto borrado con exito')
            return
    print('Contacto no encontrado')

def save_load_contacts():
    global agenda_contactos
    print('\nHas elegido la opción 5: Guardar y Cargar Contactos')
    
    # Guardar contactos
    print('Guardando contactos...')
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.append(["Nombre", "Telefono", "Correo"])  # Encabezados
    for contacto in agenda_contactos:
        hoja.append([contacto['Nombre'], contacto['Telefono'], contacto['Correo']])
    
    try:
        wb.save(nombre_archivo)
        print(f"Contactos guardados exitosamente en {nombre_archivo}")
    except Exception as e:
        print(f"Error al guardar los contactos: {e}")
        return

    # Cargar contactos para verificar
    print('Cargando contactos para verificar...')
    agenda_contactos = []  # Reiniciar la lista de contactos
    
    try:
        wb = openpyxl.load_workbook(nombre_archivo)
        hoja = wb.active
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            nombre, telefono, correo = fila
            contacto = {
                "Nombre": nombre,
                "Telefono": str(telefono),
                "Correo": correo
            }
            agenda_contactos.append(contacto)
        print(f"Se cargaron {len(agenda_contactos)} contactos desde {nombre_archivo}")
    except Exception as e:
        print(f"Error al cargar los contactos: {e}")
    
    return agenda_contactos

menu_contactos = {
    '1': ('1. Agregar un Contacto', add_contact),
    '2': ('2. Mostrar Contactos', show_contact),
    '3': ('3. Buscar un Contacto', find_contact),
    '4': ('4. Eliminar un Contacto', delete_contact),
    '5': ('5. Guardar y Cargar Contactos', save_load_contacts),
}

#Menú de opciones, se llama a cada una de las funciones
def menu():
    while True:
        print("\nMenú de Agenda de Contactos")
        for i in sorted(menu_contactos):
            print(menu_contactos[i][0])
        
        opcion = input("Selecciona una opción (1-5) o 'x' para salir: ")
        
        if opcion == 'x':
            break
        elif opcion in menu_contactos:
            menu_contactos[opcion][1]()
        else:
            print("Opción no válida, por favor intenta de nuevo.")

menu()
